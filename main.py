from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
import openpyxl

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

wait = WebDriverWait(driver, 20)

wb = openpyxl.load_workbook("data/data.xlsx")
ws = wb['Sheet1']
rows = ws.max_row + 1
for i in range(2, rows):
    idx = ws.cell(row=i, column=1).value
    passx = ws.cell(row=i, column=2).value
    dmcx = ws.cell(row=i, column=3).value

    driver.get('SITE/login')

    userid = driver.find_element(By.ID, 'username')
    userid.send_keys(idx)

    password = driver.find_element(By.ID, 'password')
    password.send_keys(passx)

    signin = driver.find_element(By.XPATH, '//*[@id="root"]/main/div[2]/div/form/button')
    signin.click()

    wait.until(ec.title_is("Home Page"))
    driver.get("SITE/Webportal/Documents")

    wait.until(ec.element_to_be_clickable((By.XPATH, '//*[@id="DMC"]/div[1]'))).click()

    download = driver.find_elements(By.XPATH, "//div[@id='DMC-content']//div//div//div//div[1]//div[1]//div[1]//div[2]//span[2]//*[name()='svg']")
    time.sleep(1)
    m = len(download)
    o = m - dmcx

    for j in range(0, o):
        download[j].click()
        time.sleep(10)
        dmcx += 1
        ws.cell(row=i, column=3).value = dmcx

wb.save("data/data.xlsx")
